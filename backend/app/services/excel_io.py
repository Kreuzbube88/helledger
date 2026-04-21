import io
from datetime import datetime, timezone
import openpyxl
from sqlalchemy.orm import Session
from app.models.transaction import Transaction


def import_excel(content: bytes, account_id: int, household_id: int, db: Session) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    imported = 0
    errors = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            raw_date, description, amount, tx_type = row[0], row[1], row[2], row[3]
            if raw_date is None or amount is None:
                errors += 1
                continue

            if isinstance(raw_date, datetime):
                tx_date = raw_date.date()
            else:
                tx_date = datetime.strptime(str(raw_date), "%Y-%m-%d").date()

            tx_type = str(tx_type).lower() if tx_type else "expense"
            if tx_type not in ("income", "expense", "transfer"):
                tx_type = "expense"

            tx = Transaction(
                household_id=household_id,
                account_id=account_id,
                date=tx_date,
                description=str(description) if description else "",
                amount=float(amount),
                transaction_type=tx_type,
                created_at=datetime.now(timezone.utc),
                updated_at=datetime.now(timezone.utc),
            )
            db.add(tx)
            imported += 1
        except Exception:
            errors += 1
            continue

    if imported > 0:
        db.commit()

    return {"imported": imported, "duplicates": 0, "errors": errors}


def export_excel(transactions: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"
    ws.append(["date", "description", "type", "amount", "account_id", "category_id"])
    for tx in transactions:
        ws.append([
            str(tx.date),
            tx.description,
            tx.transaction_type,
            float(tx.amount),
            tx.account_id,
            tx.category_id,
        ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
